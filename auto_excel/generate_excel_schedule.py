import random
import pandas as pd
import openpyxl


def column_sum_formula_generator(max_row, row_i, col_i):
    col_letter = chr(ord('@') + col_i)
    formula = '=COUNTIF({col}$2:{col}${max_row},$A{row_i})'.format(
        col=col_letter,
        max_row=max_row,
        row_i=row_i)
    return formula


def add_column_sums(ws, rotations, blocks):
    max_row = ws.max_row
    row_i = max_row + 2
    for rotation in rotations:
        ws.cell(column=1, row=row_i, value=rotation)
        for col_i in range(2, len(blocks) + 2):
            formula = column_sum_formula_generator(max_row, row_i, col_i)
            ws.cell(column=col_i, row=row_i, value=formula)
        row_i += 1


def generate_excel_schedule(doctor_names, blocks, rotations):
    """
    :param doctor_names:
    :param blocks:
    :param rotations:
    """
    block_dict = {}
    for block in blocks:
        block_rotation_assignments = []
        for _ in doctor_names:
            block_rotation_assignments.append(random.choice(rotations))
        block_dict[block] = block_rotation_assignments
    # df index are doctor names; columns are blocks
    # values are rotation assignment
    df = pd.DataFrame(data=block_dict, index=doctor_names)
    print(df)
    df.to_excel(file_name)
    wb = openpyxl.load_workbook(filename=file_name)
    ws = wb.active
    add_column_sums(ws, rotations, blocks)
    wb.save(filename=file_name)
    wb.close()


if __name__ == '__main__':
    file_name = ''
    doctornames = []
    blocknames = []
    rotationnames = []

    generate_excel_schedule(doctornames, blocknames, rotationnames)
